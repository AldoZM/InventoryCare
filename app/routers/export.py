import tempfile
from pathlib import Path
from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from app.auth import require_admin
from app.database import get_db

router = APIRouter(prefix="/api", tags=["export"])


@router.get("/export")
def export_excel(conn=Depends(get_db), _=Depends(require_admin)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")

    def _write_sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(list(row))
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    cur = conn.cursor()

    # Products
    ws = wb.active
    ws.title = "Productos"
    cur.execute("SELECT id, sku, name, description, unit, min_stock, created_at FROM products ORDER BY id")
    _write_sheet(ws, ["ID", "SKU", "Nombre", "Descripción", "Unidad", "Stock Mín.", "Creado"], cur.fetchall())

    # Locations
    ws2 = wb.create_sheet("Ubicaciones")
    cur.execute("SELECT id, code, name, description FROM locations ORDER BY id")
    _write_sheet(ws2, ["ID", "Código", "Nombre", "Descripción"], cur.fetchall())

    # Inventory
    ws3 = wb.create_sheet("Inventario")
    cur.execute("""
        SELECT p.sku, p.name, l.code, l.name, i.quantity
        FROM inventory i
        JOIN products p ON p.id = i.product_id
        JOIN locations l ON l.id = i.location_id
        ORDER BY p.name, l.code
    """)
    _write_sheet(ws3, ["SKU", "Producto", "Cód. Ubicación", "Ubicación", "Cantidad"], cur.fetchall())

    # Movements
    ws4 = wb.create_sheet("Movimientos")
    cur.execute("""
        SELECT m.id, p.sku, p.name, l.code, m.type, m.quantity,
               m.reference, m.notes, u.username, m.created_at
        FROM movements m
        JOIN products p ON p.id = m.product_id
        JOIN locations l ON l.id = m.location_id
        LEFT JOIN users u ON u.id = m.user_id
        ORDER BY m.created_at DESC
        LIMIT 5000
    """)
    _write_sheet(ws4, ["ID", "SKU", "Producto", "Ubicación", "Tipo", "Cantidad", "Referencia", "Notas", "Usuario", "Fecha"], cur.fetchall())

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = Path(tmp.name)

    wb.save(str(tmp_path))

    return FileResponse(
        path=str(tmp_path),
        filename="inventarycare_export.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
